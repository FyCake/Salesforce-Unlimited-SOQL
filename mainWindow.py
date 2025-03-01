from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, QFileDialog, QProgressBar, QTextEdit, QMessageBox
from PyQt5.QtCore import  pyqtSignal, QProcess
from openpyxl import load_workbook
import os
from query import QueryThread

class SalesforceQueryApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Salesforce Unlimited SOQL")
        self.setGeometry(100, 100, 1000, 2000)
        self.is_query_running = False  # 新增标志变量，用于跟踪查询状态
        self.error_signal = pyqtSignal(str)  # 确保 error_signal 被正确初始化
        
        self.process = QProcess()
        self.initUI()
        self.log_text.append(f"❤❤❤\n> 使用须知：\n· 1.使用前请安装Salesforce CLI，如果不会配置Path，请安装在默认路径即可。Salesforce CLI下载地址：https://developer.salesforce.com/tools/salesforcecli\n· 2.Org Alias Name 为您连接Salesforce Org时自定义的名称，用于区分不同的Org；如第一次该工具，点击查询按钮后，程序会跳转到Salesforce登录页面，进行登录并连接Org，非第一次使用，既已经登录验证过了，可直接输入Org Alias Name拉取数据。\n· 3.程序会读取Excel文件中名为'Sheet1'子表包含标题的第一列数据，作为SOQL查询条件，查询前程序会自动去重，去重后的数据会保存到名为'Sheet1去重后list'的子表中。请确保Excel文件中'Sheet1'子表的第一列数据为包含标题的查询条件。\n· 4.SOQL语句须为合法的SOQL语句，如需查询所有数据，请使用'SELECT FIELD1, FIELD2, FIELD3 FROM object_name WHERE FIELD'语句形式，程序中会自动添加IN和'()'，你不用自己添加，只需要确保Sheet1作为查询条件的数据字段名在SOQL语句的最后即可。\n· 5.查询结果会保存到名为'SOQL Result'的子表中。\n❤❤❤\n")
        
        self.credentials = None  # 新增成员变量来存储凭证

    def initUI(self):
        # 主布局
        main_layout = QVBoxLayout()

        # Org alias 输入,默认为fy163，用户可以更改
        self.alias_label = QLabel("Org Alias Name:")
        self.alias_input = QLineEdit(text="fy163")
        main_layout.addWidget(self.alias_label)
        main_layout.addWidget(self.alias_input)

        # Excel 文件路径输入
        self.file_label = QLabel("Excel 文件路径:")
        self.file_input = QLineEdit(text="C:/Users/FY/Desktop/AI项目/PythonCode/workpy/testData.xlsx")
        self.file_btn = QPushButton("选择文件")
        self.file_btn.clicked.connect(self.select_file)
        file_layout = QHBoxLayout()
        file_layout.addWidget(self.file_input)
        file_layout.addWidget(self.file_btn)
        main_layout.addWidget(self.file_label)
        main_layout.addLayout(file_layout)

        # SOQL 查询语句输入
        self.soql_label = QLabel("SOQL 查询语句:")
        self.soql_input = QLineEdit(text="select id,name from account where id")
        main_layout.addWidget(self.soql_label)
        main_layout.addWidget(self.soql_input)

        # 查询按钮，一旦被点击，按钮就不可用，直到查询结束
        self.query_btn = QPushButton("执行查询")
        self.query_btn.clicked.connect(self.execute_query)
        main_layout.addWidget(self.query_btn)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        # log 输出标签
        self.log_label = QLabel("日志输出:")
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)  # 设置为只读
        main_layout.addWidget(self.log_label)
        main_layout.addWidget(self.log_text)

        # 输出结果标签
        # self.result_label = QLabel("查询结果会显示在下面...")
        # main_layout.addWidget(self.result_label)

        self.setLayout(main_layout)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel Files (*.xlsx)")
        if file_path:
            self.file_input.setText(file_path)
    def pre_check(self, alias, soql_query_input, file_path, sheet_name):
        if not alias or not file_path or not soql_query_input:
            self.error_stop("Org Alias Name、Excel 文件路径、SOQL 查询语句 均不能为空！")
            return
        
        # 检查 Excel 文件是否存在
        if not os.path.exists(file_path):
            self.error_stop("Excel 文件不存在！请重新选择路径！")
            return

        # 检查 Excel 文件是否被用户打开了，尝试用写入模式打开
        try:
            with open(file_path, 'a') as f:
                pass
        except Exception as e:
            self.error_stop(f"Excel 文件被打开！请关闭文件后重试！")
            return
        
        # 检查 Excel 文件是否有 'Sheet1' 子表
        try:
            workbook = load_workbook(file_path)
            if sheet_name not in workbook.sheetnames:
                self.error_stop("Excel 文件中没有 'Sheet1' 子表！请检查文件内容！")
                return
        except Exception as e:
            self.error_stop(f"Excel 文件打开失败！请检查文件内容！")
            return
        return True
    
    def execute_query(self):
        if self.is_query_running:
            return
        # 设置查询状态为进行中，并禁用查询按钮
        self.is_query_running = True
        self.query_btn.setEnabled(False)

        # 进度条置零
        self.progress_bar.setValue(0)

        alias = self.alias_input.text().strip()
        file_path = self.file_input.text().strip()
        soql_query_input = self.soql_input.text().strip() 

        # 文件检查
        pre_check_result = self.pre_check(alias, soql_query_input, file_path, 'Sheet1')
        if not pre_check_result:
            return

        self.query_thread = QueryThread(alias, file_path, soql_query_input)
        
        # 连接信号
        self.query_thread.progress_signal.connect(self.update_progress)
        self.query_thread.log_signal.connect(self.log_text.append)
        self.query_thread.error_signal.connect(self.error_stop)

        self.query_thread.start()
        self.query_thread.finished.connect(self.on_query_finished)

    def on_query_finished(self):
        self.is_query_running = False
        self.query_btn.setEnabled(True)

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)
    
    def error_stop(self, error_msg):
        QMessageBox.warning(self, "Error", error_msg)
        self.is_query_running = False
        self.query_btn.setEnabled(True)
    