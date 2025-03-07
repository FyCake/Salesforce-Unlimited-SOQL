from PyQt5.QtCore import QThread, pyqtSignal
from simple_salesforce import Salesforce
import json
import os
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from collections import deque
import time

class QueryThread(QThread):
    progress_signal = pyqtSignal(int)
    error_signal = pyqtSignal(str)  # 错误信号,并释放button
    log_signal = pyqtSignal(str)  # 日志信号


    def __init__(self, alias, file_path, soql_query_input, batch_size_input):
        super().__init__()
        self.alias = alias
        self.file_path = file_path
        self.soql_query_input = soql_query_input
        self.batch_size_input = batch_size_input
        self.is_running = True

    def run(self):
        self.log_signal.emit(f"*********\n> 查询开始时间：{time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))}")
        # 获取 Salesforce 凭证
        self.credentials = self.get_sf_credentials(self.alias)
        if not self.credentials:
            self.log_signal.emit("❗❗❗ 无法获取 Salesforce 凭证。请检查您的 Salesforce CLI 配置！\n*********\n")
            self.error_signal.emit("无法获取 Salesforce 凭证。请检查您的 Salesforce CLI 配置！")
            self.is_running = False
            return
        if self.is_running:
            self.pre_query(self.soql_query_input, self.file_path, 'Sheet1', self.credentials)
            if self.is_running:
                self.query_account(self.credentials, self.batch_size_input)
                self.log_signal.emit("⭐⭐⭐查询完成，数据已导出！⭐⭐⭐\n*********\n")             

    def get_sf_credentials(self, alias):
        # 获取凭证，返回 session_id, instance_url，如果无法获取凭证，则跳转到登录页面
        try:
            self.log_signal.emit(f"> 获取 {alias} 组织凭证...")
            sf_info = os.popen(f"sfdx force:org:display -u {alias} --json").read()#这句话是获取org信息，包括accessToken和instanceUrl
            sf_info_json = json.loads(sf_info)
            if 'result' not in sf_info_json or 'accessToken' not in sf_info_json['result'] or 'instanceUrl' not in sf_info_json['result']:
                self.log_signal.emit(f"> 请在浏览器中登录并授权访问 {alias} 组织...")
                os.system(f"sfdx force:auth:web:login -a {alias}")
                sf_info = os.popen(f"sfdx force:org:display -u {alias} --json").read()
                sf_info_json = json.loads(sf_info)
                self.credentials = self.get_sf_credentials(alias)
                if not self.credentials:
                    self.log_signal.emit("❗❗❗ 仍然无法获取 Salesforce 凭证。请手动检查您的 Salesforce CLI 配置！\n*********\n")
                    self.error_signal.emit("仍然无法获取 Salesforce 凭证。请手动检查您的 Salesforce CLI 配置！")
                    self.is_running = False
                    return None
        except Exception as e:
            return None
        org_data = sf_info_json['result']
        session_id = org_data['accessToken']
        instance_url = org_data['instanceUrl']

        return session_id, instance_url

    def sanitize_soql_input(self, user_input):
        # user_input = re.sub(r"[^\w\s,']", "", user_input)
        return user_input
    
    def pre_query(self, soql_query_input, file_path, sheet_name, credentials=None):
        # 先为file中ids去重
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        ids_header = df.iloc[0, 0]
        df = df.iloc[1:, 0]
        df = df.drop_duplicates()
        ids = df.tolist()
        ids = list(set(ids))
        drop_list = [ids_header] + ids
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            df = pd.DataFrame(drop_list)
            df.to_excel(writer, sheet_name='Sheet1去重后list', startrow=0, index=False, header=False)
        if ids_header is None:
            self.log_signal.emit(f"> 去重执行完毕！{file_path.split('/')[-1]}的Sheet1去重后共有{len(ids)}个待查询数据。")
        else:
            self.log_signal.emit(f"> 去重执行完毕！{file_path.split('/')[-1]}的Sheet1去重后共有{len(ids)}个待查询{ids_header}。")

        # 在从子表 Sheet1去重后list 中取出第一个id，作为soql查询条件，用于检测整个查询语句是否合法
        first_id = ids[0]
        soql_query_input_test = soql_query_input + " IN ('{}')".format(first_id)

        session_id, instance_url = credentials
        # 检查 SOQL 查询语句是否合法
        sf = Salesforce(session_id=session_id, instance_url=instance_url)
        try:
            results = sf.query(soql_query_input_test)
            if 'errorCode' in results:
                self.log_signal.emit(f"❗❗❗ SOQL查询语句不合法！\n· 请检查你的SOQL语句: {soql_query_input}\n错误原因：{results['errorCode']}\n*********\n")
                self.error_signal.emit(f"SOQL 查询语句不合法！请检查你的SOQL语句是否正确!\n")
                self.is_running = False
                return
        except Exception as e:
            self.log_signal.emit(f"❗❗❗ SOQL查询语句不合法！\n· 请检查你的SOQL语句: {soql_query_input}\n错误原因：{e}\n*********\n")
            self.error_signal.emit(f"SOQL 查询语句不合法！请检查你的SOQL语句是否正确!\n")
            self.is_running = False
            return

    def query_account(self, credentials, batch_size_input):
        session_id, instance_url = credentials
        sf = Salesforce(instance_url=instance_url, session_id=session_id)
        if not sf:
            return

        # 清理 SOQL 输入
        soql_query_input_safe = self.sanitize_soql_input(self.soql_query_input)
        df = pd.read_excel(self.file_path, sheet_name='Sheet1去重后list')
        ids = df.iloc[:, 0].tolist()
        
        batch_size_input = int(batch_size_input)
        all_results = []
        self.log_signal.emit(f"> 开始 SOQL 查询数据...")
        for i in range(0, len(ids), batch_size_input):#这句话是把ids分成batch_size_input个子列表
            self.progress_signal.emit(int((i + batch_size_input - 2) / len(ids) * 100))
            if len(ids) - i < batch_size_input:
                self.log_signal.emit(f"· 正在查询第{i+1}到{len(ids)}条数据...")
            elif i > len(ids) - batch_size_input:
                self.log_signal.emit(f"· 正在查询第{i+1}到{len(ids)}条数据...")
            else:
                self.log_signal.emit(f"· 正在查询第{i+1}到{i+batch_size_input}条数据...")
            batch_ids = ids[i:i + batch_size_input]
            soql_query = soql_query_input_safe + " IN ({})".format(','.join(map(lambda x: "'{}'".format(x), batch_ids)))
            try:
                results = sf.query(soql_query)
            except Exception as e:
                ids_header = df.iloc[0, 0]
                self.log_signal.emit(f"❗❗❗ Salesforce SOQL查询错误，请检查你Sheet1第一列所需查询{ids_header} List是否正确！\n*********\n")
                self.error_signal.emit(f"Salesforce SOQL查询错误，请检查你Sheet1第一列所需查询{ids_header} List是否正确！")
                self.is_running = False
                return
            all_results.extend(results['records'])
        self.log_signal.emit(f"> SOQL 查询执行完毕！")

        # 如果all_results为空，则直接返回
        if not all_results:
            self.log_signal.emit(f"❗❗❗ 查询完毕，SOQL查询结果没有数据返回！\n*********\n")
            self.is_running = False
            return

        self.log_signal.emit(f"> 开始导出数据...")
        expanded_data = [self.expand_dict_iterative(item) for item in all_results]

        select_fields = soql_query_input_safe.lower().split('select')[1].split('from')[0].strip().split(',')
        select_fields = [field.strip() for field in select_fields]

        # 检查是否有包含address的字段
        has_address_field = any('address' in field.lower() for field in select_fields)

        # 从 all_results 中取出一个 item 用于筛选所需字段
        sample_item = expanded_data[0] if expanded_data else {}

        # 筛选出所需字段
        filtered_fields = [k for k in sample_item.keys() if any(k.lower() == field.lower() for field in select_fields) or (has_address_field and 'address' in k.lower())]

        # 根据筛选出的字段，从 all_results 中清理所需数据
        expanded_data = [
            {k: v for k, v in item.items() if k in filtered_fields}
            for item in expanded_data
        ]

        results_df = pd.DataFrame(expanded_data)

        # 检查是否已经存在SOQL Result子表，如果存在，则存入SOQL Result1、SOQL Result2...等子表中
        i = 1
        if self.check_sheet_exists(self.file_path, 'SOQL Result'):
            for j in range(i, 100):
                if self.check_sheet_exists(self.file_path, f'SOQL Result{j}'):
                    i += 1
                if i == 100:
                    self.log_signal.emit(f"❗❗❗ 超过100个SOQL Result子表，请删除旧的子表后再运行查询！\n*********\n")
                    self.error_signal.emit(f"超过100个SOQL Result子表，请删除旧的子表后再运行查询！")
                    self.is_running = False
                    return
            result_sheet_name = f'SOQL Result{i}'
        else:
            result_sheet_name = 'SOQL Result'

        with pd.ExcelWriter(self.file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name=result_sheet_name, index=False)
            worksheet = writer.sheets[result_sheet_name]
            for col in worksheet.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left')
        self.progress_signal.emit(100)
        self.log_signal.emit(f"> 查询结果已保存到{self.file_path.split('/')[-1]}的{result_sheet_name}子表中。")

    def expand_dict_iterative(self, data, sep='.', remove_empty=True, remove_null=True):
        stack = deque([(data, '')])
        new_dict = {}

        while stack:
            current, parent_key = stack.pop()
            if isinstance(current, dict):
                for k, v in current.items():
                    new_key = f"{parent_key}{sep}{k}" if parent_key else k
                    if isinstance(v, dict):
                        stack.append((v, new_key))
                    elif isinstance(v, list):
                        for idx, item in enumerate(v):
                            if isinstance(item, dict):
                                stack.append((item, f"{new_key}[{idx}]"))
                            else:
                                new_dict[f"{new_key}[{idx}]"] = item
                    else:
                        new_dict[new_key] = v
            else:
                new_dict[parent_key] = current
        return new_dict

    def check_sheet_exists(self, file_path, sheet_name):
        try:
            wb = load_workbook(file_path)
            return sheet_name in wb.sheetnames
        except Exception as e:
            return False
